import os, re, sys, glob, subprocess, time
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ===== CONFIG =====
# Sistema de busca inteligente de diretório
def encontrar_base_dir():
    """
    Procura o diretório base onde o programa está rodando.
    Funciona em qualquer pasta, qualquer PC.
    """
    # Modo executável (PyInstaller) - usa pasta onde o .exe está
    if getattr(sys, 'frozen', False):
        exe_dir = os.path.dirname(sys.executable)
        print(f"✅ Usando pasta do executável: {exe_dir}")
        return exe_dir
    else:
        # Script Python (desenvolvimento) - usa pasta do script
        script_dir = os.path.dirname(os.path.abspath(__file__))
        # Sobe um nível se estiver em /scripts
        if os.path.basename(script_dir) == 'scripts':
            script_dir = os.path.dirname(script_dir)
        print(f"✅ Usando pasta do script: {script_dir}")
        return script_dir

BASE_DIR   = encontrar_base_dir()
FILHAS_DIR = os.path.join(BASE_DIR, "filhas")
MAE_PATH   = os.path.join(BASE_DIR, "PLANILHA_MAE.xlsx")

NOME_REGEX = re.compile(r"^([A-ZÇÃÕÉÊÁÍÓÚ]+)_[A-ZÇÃÕÉÊÁÍÓÚ]+ - ATENDIMENTOS.{0,3}(\d{2}-\d{2}-\d{2}|\d{2}-\d{2}-\d{4})\.xlsm?$", re.I)
COLS_ESPERADAS = [
    "INICIAR","RESPONSÁVEL","OPERAÇÃO","CLIENTE","SOLICITAÇÃO","SETOR",
    "OBSERVAÇÕES","FINALIZAR","TIME SPENT","TRATATIVA SETOR","TIME SPENT - SETOR"
]

def fechar_excel():
    """Verifica se o Excel está aberto e pergunta se quer fechar"""
    try:
        print("🔄 Verificando se o Excel está aberto...")
        result = subprocess.run(['tasklist', '/FI', 'IMAGENAME eq EXCEL.EXE'], 
                              capture_output=True, text=True, shell=True)
        
        if 'EXCEL.EXE' in result.stdout:
            print("\n⚠️  ATENÇÃO: Excel está aberto!")
            print("⚠️  Para evitar erros, é recomendado fechar o Excel antes de continuar.")
            print("⚠️  Todas as planilhas abertas serão fechadas (SALVE SEU TRABALHO!).")
            print("\n💡 OPÇÕES:")
            print("   1. Fechar Excel automaticamente")
            print("   2. Criar arquivo temporário (abre em nova janela do Excel)")
            print("   3. Cancelar operação")
            
            resposta = input("\n🤔 Escolha uma opção (1/2/3): ").strip()
            
            if resposta == '1':
                print("📊 Fechando Excel...")
                subprocess.run(['taskkill', '/F', '/IM', 'EXCEL.EXE'], 
                             capture_output=True, shell=True)
                time.sleep(2)  # Aguarda o Excel fechar completamente
                print("✅ Excel fechado com sucesso!")
                return 'fechado'
            elif resposta == '2':
                print("📋 Modo arquivo temporário selecionado!")
                return 'navegador'
            else:
                print("❌ Operação cancelada pelo usuário.")
                return 'cancelado'
        else:
            print("✅ Excel não está aberto.")
            return 'livre'
    except Exception as e:
        print(f"⚠️ Erro ao verificar Excel: {e}")
        return 'livre'

def abrir_planilha_final():
    """Abre a planilha final automaticamente após o processamento"""
    try:
        if not os.path.exists(MAE_PATH):
            print("⚠️ Arquivo final não encontrado para abrir.")
            return False
        
        # Pergunta ao usuário se quer abrir
        print("\n" + "="*60)
        print("📊 Planilha consolidada com sucesso!")
        print(f"📁 Local: {MAE_PATH}")
        print("="*60)
        
        resposta = input("\n🤔 Deseja abrir a planilha agora? (S/N): ").strip().upper()
        
        if resposta in ['S', 'SIM', 'Y', 'YES']:
            print(f"📊 Abrindo planilha: {os.path.basename(MAE_PATH)}")
            # Usa os.startfile() que é mais confiável no Windows
            os.startfile(MAE_PATH)
            print("✅ Planilha aberta!")
            return True
        else:
            print("⏭️ Planilha não foi aberta. Você pode abri-la manualmente depois.")
            return False
            
    except Exception as e:
        print(f"⚠️ Erro ao abrir planilha final: {e}")
        print(f"💡 Abra manualmente em: {MAE_PATH}")
        return False

def abrir_no_navegador(df: pd.DataFrame):
    """Salva como arquivo temporário para visualização paralela"""
    try:
        # Limpa arquivos temporários antigos primeiro
        print(f"\n🧹 Limpando arquivos temporários antigos...")
        temp_files = glob.glob(os.path.join(BASE_DIR, "PLANILHA_TEMP_*.xlsx"))
        if temp_files:
            for temp_file in temp_files:
                try:
                    os.remove(temp_file)
                    print(f"   🗑️ Removido: {os.path.basename(temp_file)}")
                except Exception as e:
                    print(f"   ⚠️ Não foi possível remover {os.path.basename(temp_file)}: {e}")
            print(f"✅ Limpeza concluída!")
        else:
            print(f"✅ Nenhum arquivo temporário antigo encontrado.")
        
        # Cria arquivo temporário com timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        temp_path = os.path.join(BASE_DIR, f"PLANILHA_TEMP_{timestamp}.xlsx")
        
        print(f"\n📊 Salvando arquivo temporário: {os.path.basename(temp_path)}")
        
        # Usa a mesma função de salvar_no_excel, mas com caminho temporário
        # Temporariamente substitui MAE_PATH
        global MAE_PATH
        mae_path_original = MAE_PATH
        MAE_PATH = temp_path
        
        try:
            salvar_no_excel(df)
            print(f"✅ Arquivo temporário salvo com TODAS as abas!")
            print(f"📁 Local: {temp_path}")
            print(f"\n💡 IMPORTANTE: Este é um arquivo temporário para visualização.")
            print(f"💡 Ele será automaticamente removido na próxima execução.")
            print(f"💡 Feche o Excel principal e execute novamente para atualizar o arquivo definitivo.")
        finally:
            # Restaura o caminho original
            MAE_PATH = mae_path_original
        
        # Pergunta se quer abrir o temporário
        resposta = input("\n🤔 Deseja abrir o arquivo temporário agora? (S/N): ").strip().upper()
        
        if resposta in ['S', 'SIM', 'Y', 'YES']:
            os.startfile(temp_path)
            print("✅ Arquivo temporário aberto em nova janela do Excel!")
        else:
            print("⏭️ Arquivo temporário salvo, mas não foi aberto.")
        
        return True
        
    except Exception as e:
        print(f"❌ Erro ao criar arquivo temporário: {e}")
        return False

def escrever_status(mensagem: str):
    try:
        wb = load_workbook(MAE_PATH)
        ws = wb.active  # primeira aba (pode ajustar pelo nome)
        ws["A1"] = mensagem
        wb.save(MAE_PATH)
    except Exception:
        pass  # se ainda não existe, segue

def aplicar_formatacao(ws, df):
    """Aplica formatação profissional na planilha"""
    
    # Estilo do cabeçalho
    header_font = Font(bold=True, color="FFFFFF")  # Cor do texto (branco)
    header_fill = PatternFill(start_color="ed7d31", end_color="ed7d31", fill_type="solid")  # Cor de fundo (laranja)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Bordas na cor laranja para todas as células
    border_laranja = Border(
        left=Side(style='thin', color='ed7d31'), 
        right=Side(style='thin', color='ed7d31'), 
        top=Side(style='thin', color='ed7d31'), 
        bottom=Side(style='thin', color='ed7d31')
    )
    
    # Aplica formatação no cabeçalho
    for col_idx, column in enumerate(ws.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_laranja
    
    # Autoajuste das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Ajusta largura com limite máximo
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Formatação das células de dados
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border_laranja
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Formatação específica para colunas de data/hora
    date_columns = ["INICIAR", "FINALIZAR", "TRATATIVA SETOR"]
    time_columns = ["TIME SPENT", "TIME SPENT - SETOR"]
    
    for col_idx, col_name in enumerate(df.columns, 1):
        if col_name in date_columns:
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                # Formatação de data bonita: DD/MM/AAAA HH:MM (sem dia da semana)
                cell.number_format = 'DD/MM/YYYY HH:MM'
        elif col_name in time_columns:
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                # Formatação de tempo: HH:MM:SS
                cell.number_format = 'HH:MM:SS'

def validar_nome(arquivo):
    nome = os.path.basename(arquivo)
    return bool(NOME_REGEX.match(nome))

def extrair_primeiro_nome(arquivo):
    nome = os.path.basename(arquivo)
    m = NOME_REGEX.match(nome)
    if not m: return None, None
    primeiro = m.group(1)
    data_txt = m.group(2)
    # Normaliza data
    try:
        if len(data_txt.split("-")[-1]) == 2:
            data = datetime.strptime(data_txt, "%d-%m-%y").date()
        else:
            data = datetime.strptime(data_txt, "%d-%m-%Y").date()
    except:
        data = None
    return primeiro.upper(), data

def ler_filhos():
    arquivos = sorted(glob.glob(os.path.join(FILHAS_DIR, "*.xlsx")) + 
                      glob.glob(os.path.join(FILHAS_DIR, "*.xlsm")))
    logs = []
    dfs = []
    for arq in arquivos:
        if not validar_nome(arq):
            logs.append(f"❌ Nome inválido: {os.path.basename(arq)} (padrão: NOME_SOBRENOME - ATENDIMENTOS - DD-MM-AA.xlsx)")
            continue
        try:
            # Lê Excel preservando formatação e quebras de linha
            df = pd.read_excel(arq, dtype=str, keep_default_na=False)
            # Remove espaços extras dos nomes das colunas
            df.columns = df.columns.str.strip()
            # Substitui células vazias por NaN para processamento correto
            df = df.replace('', pd.NA)
            faltantes = [c for c in COLS_ESPERADAS if c not in df.columns]
            if faltantes:
                logs.append(f"❌ Colunas faltando em {os.path.basename(arq)}: {faltantes}")
                continue
            
            # FILTRAR LINHAS VAZIAS - Remove linhas onde todas as colunas principais estão vazias
            colunas_principais = ["RESPONSÁVEL", "OPERAÇÃO", "CLIENTE", "SOLICITAÇÃO"]
            
            # Verifica se pelo menos uma coluna principal tem dados válidos
            mask_nao_vazio = False
            for col in colunas_principais:
                if col in df.columns:
                    # Verifica se a coluna tem valores não-nulos e não-vazios (preservando quebras de linha)
                    mask_col = df[col].notna() & (df[col] != '') & (df[col].astype(str).str.strip() != "")
                    if mask_nao_vazio is False:
                        mask_nao_vazio = mask_col
                    else:
                        mask_nao_vazio = mask_nao_vazio | mask_col
            
            # Se não encontrou nenhuma linha válida, usar máscara padrão
            if mask_nao_vazio is False:
                mask_nao_vazio = df.index >= 0  # Todas as linhas
            
            df_filtrado = df[mask_nao_vazio].copy()
            
            if df_filtrado.empty:
                logs.append(f"⚠️ Arquivo sem dados válidos: {os.path.basename(arq)}")
                continue
            
            primeiro, data_arq = extrair_primeiro_nome(arq)
            df_filtrado["PRIMEIRO_NOME"] = primeiro
            df_filtrado["DATA_ARQUIVO"]  = pd.to_datetime(data_arq) if data_arq else pd.NaT
            df_filtrado["ARQUIVO"]       = os.path.basename(arq)
            dfs.append(df_filtrado[COLS_ESPERADAS + ["PRIMEIRO_NOME","DATA_ARQUIVO","ARQUIVO"]])
            logs.append(f"✅ OK: {os.path.basename(arq)} ({len(df_filtrado)} linhas úteis de {len(df)} total)")

        except Exception as e:
            logs.append(f"❌ Erro lendo {os.path.basename(arq)}: {e}")
    return (pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame(columns=COLS_ESPERADAS+["PRIMEIRO_NOME","DATA_ARQUIVO","ARQUIVO"])), logs

def salvar_no_excel(df: pd.DataFrame):
    # Cria pasta de backup se não existir
    backup_dir = os.path.join(BASE_DIR, "backup")
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir)
        print(f"📁 Pasta de backup criada: {backup_dir}")
    
    # Faz backup da planilha mãe antes de apagar (se existir)
    if os.path.exists(MAE_PATH):
        try:
            backup_path = os.path.join(backup_dir, "PLANILHA_MAE_BACKUP.xlsx")
            
            # Se já existe um backup, remove ele (mantém sempre só o último)
            if os.path.exists(backup_path):
                os.remove(backup_path)
                print(f"🔄 Backup anterior removido")
            
            # Cria novo backup
            import shutil
            shutil.copy2(MAE_PATH, backup_path)
            print(f"💾 Backup criado: PLANILHA_MAE_BACKUP.xlsx")
            
            # Só remove o arquivo original após backup bem-sucedido
            os.remove(MAE_PATH)
            print(f"🗑️ Arquivo anterior removido: {os.path.basename(MAE_PATH)}")
            
        except Exception as e:
            print(f"⚠️ Erro ao fazer backup: {e}")
            print(f"⚠️ Continuando sem remover arquivo anterior...")
    
    # Cria ou abre a Mãe
    if os.path.exists(MAE_PATH):
        wb = load_workbook(MAE_PATH)
    else:
        from openpyxl import Workbook
        wb = Workbook()
        # Remove a aba padrão "Sheet" que vem com workbooks novos
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    # 1) COMPILE GERAL
    aba_compilado = "COMPILE GERAL"
    if aba_compilado in wb.sheetnames:
        ws = wb[aba_compilado]
        wb.remove(ws)
    ws = wb.create_sheet(aba_compilado)

    # cabeçalho
    cols = df.columns.tolist()
    ws.append(cols)
    
    # Adiciona dados preservando quebras de linha
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if pd.notna(value) and value != '':
                # Preserva quebras de linha e formatação especial
                if isinstance(value, str):
                    # Preserva quebras de linha originais do Excel
                    cell.value = value
                    # Se tem múltiplas linhas, aplica formatação especial
                    if '\n' in str(value) or len(str(value).split()) > 5:
                        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
                    else:
                        cell.alignment = Alignment(vertical="center", horizontal="left")
                else:
                    cell.value = value
    
    # Aplica formatação
    aplicar_formatacao(ws, df)

    # 2) MÉTRICAS
    aba_metricas = "MÉTRICAS"
    if aba_metricas in wb.sheetnames:
        wb.remove(wb[aba_metricas])
    wsM = wb.create_sheet(aba_metricas)

    # KPIs básicos
    total = len(df)
    finalizados = int((df["FINALIZAR"].astype(str).str.lower().isin(["sim","yes","true","1"])).sum()) if not df.empty else 0
    tempo_total = pd.to_numeric(df["TIME SPENT"], errors="coerce").fillna(0).sum() if "TIME SPENT" in df else 0
    tempo_medio = pd.to_numeric(df["TIME SPENT"], errors="coerce").fillna(0).mean() if total else 0

    # Estilo para a aba de métricas
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="ed7d31", end_color="ed7d31", fill_type="solid")
    border_laranja = Border(
        left=Side(style='thin', color='ed7d31'), 
        right=Side(style='thin', color='ed7d31'), 
        top=Side(style='thin', color='ed7d31'), 
        bottom=Side(style='thin', color='ed7d31')
    )
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Cabeçalho principal das métricas
    wsM["A1"] = "📊 MÉTRICAS GERAIS"
    wsM.merge_cells('A1:B1')
    wsM["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    wsM["A1"].fill = header_fill
    wsM["A1"].alignment = center_alignment
    wsM["A1"].border = border_laranja

    # Cabeçalhos das colunas
    wsM["A3"] = "Indicador"
    wsM["B3"] = "Valor"
    wsM["A3"].font = header_font
    wsM["B3"].font = header_font
    wsM["A3"].fill = header_fill
    wsM["B3"].fill = header_fill
    wsM["A3"].alignment = center_alignment
    wsM["B3"].alignment = center_alignment
    wsM["A3"].border = border_laranja
    wsM["B3"].border = border_laranja
    
    # Dados das métricas
    metricas_dados = [
        ["📈 Atendimentos Totais", total],
        ["✅ Finalizados", finalizados],
        ["📊 % Finalizados", f"{(finalizados/total*100 if total else 0):.1f}%"],
        ["⏱️ Tempo Total (minutos)", f"{tempo_total:.1f}"],
        ["⏰ Tempo Médio (minutos)", f"{tempo_medio:.1f}"]
    ]
    
    row = 4
    for indicador, valor in metricas_dados:
        wsM[f"A{row}"] = indicador
        wsM[f"B{row}"] = valor
        wsM[f"A{row}"].border = border_laranja
        wsM[f"B{row}"].border = border_laranja
        wsM[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
        wsM[f"B{row}"].alignment = center_alignment
        row += 1

    # Tabelas pivôs estilizadas
    def write_pivot_styled(title, series_counts, start_row, col1_name="Item", col2_name="Quantidade"):
        # Título da tabela
        wsM[f"A{start_row}"] = f"📋 {title}"
        wsM.merge_cells(f'A{start_row}:B{start_row}')
        wsM[f"A{start_row}"].font = Font(bold=True, size=12, color="FFFFFF")
        wsM[f"A{start_row}"].fill = header_fill
        wsM[f"A{start_row}"].alignment = center_alignment
        wsM[f"A{start_row}"].border = border_laranja
        
        # Cabeçalhos da tabela com nomes específicos
        wsM[f"A{start_row+1}"] = col1_name
        wsM[f"B{start_row+1}"] = col2_name
        wsM[f"A{start_row+1}"].font = header_font
        wsM[f"B{start_row+1}"].font = header_font
        wsM[f"A{start_row+1}"].fill = header_fill
        wsM[f"B{start_row+1}"].fill = header_fill
        wsM[f"A{start_row+1}"].alignment = center_alignment
        wsM[f"B{start_row+1}"].alignment = center_alignment
        wsM[f"A{start_row+1}"].border = border_laranja
        wsM[f"B{start_row+1}"].border = border_laranja
        
        # Dados da tabela
        r = start_row + 2
        for k, v in series_counts.items():
            wsM[f"A{r}"] = str(k)
            wsM[f"B{r}"] = int(v)
            wsM[f"A{r}"].border = border_laranja
            wsM[f"B{r}"].border = border_laranja
            wsM[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center")
            wsM[f"B{r}"].alignment = center_alignment
            r += 1
        return r + 1  # Espaço extra após a tabela

    if not df.empty:
        por_setor = df["SETOR"].astype(str).value_counts()
        next_row = write_pivot_styled("ATENDIMENTOS POR SETOR", por_setor, 10, "Setor", "Qtd. Atendimentos")
        por_resp = df["PRIMEIRO_NOME"].astype(str).value_counts()
        write_pivot_styled("ATENDIMENTOS POR RESPONSÁVEL", por_resp, next_row + 1, "Responsável", "Qtd. Atendimentos")
    
    # Ajuste manual das colunas na aba de métricas
    wsM.column_dimensions['A'].width = 35
    wsM.column_dimensions['B'].width = 20

    # 3) Abas por pessoa
    if not df.empty:
        for nome in sorted(df["PRIMEIRO_NOME"].dropna().astype(str).unique()):
            sheet_name = nome.title()
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])
            wsP = wb.create_sheet(sheet_name)
            sub = df[df["PRIMEIRO_NOME"] == nome][COLS_ESPERADAS + ["DATA_ARQUIVO","ARQUIVO"]]
            
            # Remove as colunas extras das abas individuais para ficar igual ao formato mostrado
            sub_clean = df[df["PRIMEIRO_NOME"] == nome][COLS_ESPERADAS]
            
            wsP.append(sub_clean.columns.tolist())
            
            # Adiciona dados preservando quebras de linha nas abas individuais
            for row_idx, (_, row) in enumerate(sub_clean.iterrows(), start=2):
                for col_idx, value in enumerate(row, start=1):
                    cell = wsP.cell(row=row_idx, column=col_idx)
                    if pd.notna(value) and value != '':
                        # Preserva quebras de linha e formatação especial nas abas individuais
                        if isinstance(value, str):
                            cell.value = value
                            # Se tem múltiplas linhas, aplica formatação especial
                            if '\n' in str(value) or len(str(value).split()) > 5:
                                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
                            else:
                                cell.alignment = Alignment(vertical="center", horizontal="left")
                        else:
                            cell.value = value
            
            # Aplica formatação na aba individual
            aplicar_formatacao(wsP, sub_clean)

    # Deixa a primeira aba como “COMPILE GERAL”
    wb.move_sheet(wb[aba_compilado], offset=-wb.index(wb[aba_compilado]))
    wb.save(MAE_PATH)

def main():
    try:
        print("🚀 Iniciando atualização das planilhas...")
        
        # Verifica se o Excel está aberto e pede ação
        status_excel = fechar_excel()
        
        # Se usuário cancelou, encerra
        if status_excel == 'cancelado':
            print("\n❌ Operação cancelada.")
            if getattr(sys, 'frozen', False):
                input("Pressione ENTER para fechar...")
            return
        
        # Verifica se os diretórios existem
        if not os.path.exists(BASE_DIR):
            raise FileNotFoundError(f"Diretório base não encontrado: {BASE_DIR}")
        
        if not os.path.exists(FILHAS_DIR):
            raise FileNotFoundError(f"Diretório de filhas não encontrado: {FILHAS_DIR}")
        
        print(f"📁 Procurando arquivos em: {FILHAS_DIR}")
        
        escrever_status("⏳ Atualizando…")
        df, logs = ler_filhos()
        
        print(f"📊 Encontrados {len(df)} registros para consolidar")
        for log in logs:
            print(log)
        
        if df.empty:
            print("⚠️ Nenhum dado encontrado para consolidar!")
        
        # Se escolheu modo navegador, salva temporário e pula salvamento normal
        if status_excel == 'navegador':
            abrir_no_navegador(df)
            
            # Salvar log em arquivo
            log_path = os.path.join(BASE_DIR, "log_compilacao.txt")
            with open(log_path, "w", encoding="utf-8") as f:
                f.write(f"Execução em {datetime.now()} (MODO TEMPORÁRIO - Excel aberto)\n")
                f.write(f"Total de registros: {len(df)}\n")
                f.write("-" * 50 + "\n")
                f.write("\n".join(logs))
            
            print(f"\n📝 Log salvo em: {log_path}")
        else:
            # Modo normal - salva no arquivo principal
            salvar_no_excel(df)
            escrever_status(f"✅ Atualizado com sucesso — {len(df)} linhas consolidadas.")
            
            # Salvar log em arquivo
            log_path = os.path.join(BASE_DIR, "log_compilacao.txt")
            with open(log_path, "w", encoding="utf-8") as f:
                f.write(f"Execução em {datetime.now()}\n")
                f.write(f"Total de registros: {len(df)}\n")
                f.write("-" * 50 + "\n")
                f.write("\n".join(logs))
            
            print(f"✅ Processo concluído! Arquivo salvo em: {MAE_PATH}")
            print(f"📝 Log salvo em: {log_path}")
            
            # Abre a planilha final (com opção de escolha)
            abrir_planilha_final()
        
        # Pausa para ver o resultado (apenas quando executado como exe)
        if getattr(sys, 'frozen', False):
            print("\n✅ Processamento finalizado!")
            input("Pressione ENTER para fechar...")
            
    except Exception as e:
        error_msg = f"❌ Erro na atualização: {e}"
        print(error_msg)
        escrever_status(error_msg)
        
        # Salva erro no log
        try:
            log_path = os.path.join(BASE_DIR, "log_compilacao.txt")
            with open(log_path, "w", encoding="utf-8") as f:
                f.write(f"ERRO em {datetime.now()}\n")
                f.write(f"{error_msg}\n")
                f.write(f"Detalhes: {str(e)}")
        except:
            pass
        
        # Tenta abrir planilha mesmo em caso de erro (se existir)
        if os.path.exists(MAE_PATH):
            print("🔄 Tentando abrir planilha existente...")
            abrir_planilha_final()
        
        # Pausa para ver o erro (apenas quando executado como exe)
        if getattr(sys, 'frozen', False):
            print("\n❌ Erro durante o processamento!")
            input("Pressione ENTER para fechar...")
        
        raise

if __name__ == "__main__":
    main()
